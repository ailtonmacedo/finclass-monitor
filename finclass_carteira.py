from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import smtplib
import ssl
import time
from collections.abc import Iterable
from datetime import date, datetime, timedelta
from email.message import EmailMessage
from html import escape
from pathlib import Path
from typing import TypedDict, cast
from urllib.parse import urljoin

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from playwright.sync_api import (
    Browser,
    BrowserContext,
    Page,
    sync_playwright,
)
from playwright.sync_api import (
    TimeoutError as PlaywrightTimeoutError,
)

BASE_DIR = Path(__file__).resolve().parent
ENV_FILE = BASE_DIR / ".env"
load_dotenv(dotenv_path=ENV_FILE, override=False)

BASE_URL = "https://app.finclass.com/"
OUTPUT_FILE = BASE_DIR / "finclass_carteira.xlsx"
HISTORY_FILE = BASE_DIR / "finclass_historico.xlsx"
ASSETS_FILE = BASE_DIR / "finclass_ativos.txt"
SESSION_FILE = BASE_DIR / "finclass_storage_state.json"
EMAIL_STATE_FILE = BASE_DIR / "finclass_email_state.json"

HISTORY_SHEET = "Histórico"
CHANGES_SHEET = "Alterações"

# Para execução automática diária, depois que a sessão estiver validada,
# FINCLASS_HEADLESS pode ser alterado para true no .env.
HEADLESS = os.getenv(
    "FINCLASS_HEADLESS",
    "false",
).strip().lower() in {"1", "true", "yes", "on"}

DEFAULT_TIMEOUT_MS = 30_000
MANUAL_LOGIN_TIMEOUT_MS = 120_000
TABLE_STABILIZE_TIMEOUT_SECONDS = 12.0


class RecommendationTable(TypedDict):
    """Estrutura de uma categoria da carteira capturada no site."""

    category: str
    headers: list[str]
    rows: list[list[str]]


class HistoryRecord(TypedDict):
    """Um ativo dentro de um snapshot diário da carteira."""

    snapshot_date: date
    category: str
    percentage: float
    company: str
    ticker: str


class AssetRecord(TypedDict):
    """Um ativo/fundo/empresa coletado da carteira atual."""

    category: str
    name: str
    code: str


class ChangeRecord(TypedDict):
    """Uma alteração detectada entre dois snapshots."""

    change_date: date
    previous_date: date
    category: str
    company: str
    ticker: str
    change_type: str
    previous_percentage: float | None
    current_percentage: float | None


class EmailSettings(TypedDict):
    """Configuração SMTP carregada do arquivo .env."""

    host: str
    port: int
    security: str
    username: str
    password: str
    sender: str
    sender_name: str
    recipients: list[str]


# -----------------------------------------------------------------------------
# Utilidades gerais
# -----------------------------------------------------------------------------


def clean_text(value: str) -> str:
    """Normaliza espaços e preserva quebras de linha úteis."""
    lines: list[str] = []

    for line in value.replace("\xa0", " ").splitlines():
        normalized = re.sub(r"[ \t]+", " ", line).strip()

        if normalized:
            lines.append(normalized)

    return "\n".join(lines)


def safe_sheet_name(name: str, used_names: set[str]) -> str:
    """Gera um nome válido e único para uma aba do Excel."""
    cleaned = re.sub(r"[:\\/?*\[\]]", "-", name).strip()
    cleaned = cleaned[:31] or "Carteira"

    candidate = cleaned
    counter = 2

    while candidate in used_names:
        suffix = f" ({counter})"
        candidate = f"{cleaned[: 31 - len(suffix)]}{suffix}"
        counter += 1

    used_names.add(candidate)
    return candidate


def required_env(name: str) -> str:
    """Lê uma variável obrigatória do .env."""
    value = os.getenv(name)

    if value is None or not value.strip():
        raise RuntimeError(f"Variável obrigatória ausente no .env: {name}")

    return value.strip()


def env_flag(name: str, default: bool = False) -> bool:
    """Converte uma variável do .env para boolean."""
    value = os.getenv(name)

    if value is None:
        return default

    return value.strip().lower() in {
        "1",
        "true",
        "yes",
        "on",
    }


def get_credentials() -> tuple[str, str]:
    """
    Obtém as credenciais da Finclass exclusivamente do arquivo .env.

    Não solicita mais e-mail/senha pelo terminal, permitindo execução
    automática via cron ou scheduler.
    """
    if not ENV_FILE.exists():
        raise RuntimeError(
            f"Arquivo .env não encontrado em: {ENV_FILE.resolve()}"
        )

    return (
        required_env("FINCLASS_EMAIL"),
        required_env("FINCLASS_PASSWORD"),
    )


# -----------------------------------------------------------------------------
# Login / navegação
# -----------------------------------------------------------------------------


def is_login_page(page: Page) -> bool:
    """Retorna True quando o formulário de login está visível."""
    try:
        return page.locator("#email").is_visible(timeout=2_000)
    except PlaywrightTimeoutError:
        return False


def get_wallet_link(page: Page):
    """
    Localiza um link da Carteira pelo href.

    A Finclass mantém versões desktop e mobile do mesmo link no DOM.
    Preferimos o visível, mas um link apenas anexado já é suficiente
    para obter o href e navegar diretamente.
    """
    visible_link = page.locator('a[href^="/carteira/"]:visible')

    if visible_link.count() > 0:
        return visible_link.first

    return page.locator('a[href^="/carteira/"]').first


def get_wallet_url(page: Page) -> str:
    """Obtém a URL real da Carteira sem depender de clique no menu."""
    wallet_link = get_wallet_link(page)
    wallet_link.wait_for(state="attached", timeout=DEFAULT_TIMEOUT_MS)

    href = wallet_link.get_attribute("href")

    if not href:
        raise RuntimeError("O link da Carteira foi encontrado, mas não possui href.")

    return urljoin(BASE_URL, href)


def submit_login(page: Page, email: str, password: str) -> None:
    """Preenche e envia o formulário de login."""
    print("Fazendo login...")

    email_field = page.locator("#email")
    password_field = page.locator("#password")

    email_field.wait_for(state="visible")
    password_field.wait_for(state="visible")

    email_field.fill(email)
    password_field.fill(password)

    form = email_field.locator("xpath=ancestor::form[1]")
    submit = form.locator('button[type="submit"], input[type="submit"]')

    if submit.count() == 0:
        submit = page.locator(
            'button[type="submit"], '
            'input[type="submit"], '
            'button:has-text("Entrar"), '
            'button:has-text("Login")'
        )

    if submit.count() == 0:
        raise RuntimeError("Não encontrei o botão de login.")

    submit.first.click()

    wallet_link = get_wallet_link(page)

    try:
        wallet_link.wait_for(
            state="attached",
            timeout=DEFAULT_TIMEOUT_MS,
        )
    except PlaywrightTimeoutError:
        print(
            "\nO login ainda não foi confirmado.\n"
            "Se aparecer CAPTCHA, código de confirmação ou outra validação, "
            "conclua manualmente no navegador."
        )

        wallet_link.wait_for(
            state="attached",
            timeout=MANUAL_LOGIN_TIMEOUT_MS,
        )


def ensure_logged_in(
    page: Page,
    context: BrowserContext,
    email: str,
    password: str,
) -> None:
    """Abre a Finclass e autentica quando necessário."""
    page.goto(
        BASE_URL,
        wait_until="domcontentloaded",
    )

    page.wait_for_timeout(1_000)

    if is_login_page(page):
        submit_login(page, email, password)

    context.storage_state(path=str(SESSION_FILE))


def open_wallet(page: Page) -> None:
    """
    Abre a Carteira navegando diretamente para o href.

    Isso evita o erro "element is outside of the viewport" causado
    pela cópia mobile do link Carteira existente no DOM.
    """
    print("Abrindo Carteira...")

    wallet_url = get_wallet_url(page)

    page.goto(
        wallet_url,
        wait_until="domcontentloaded",
    )

    page.locator(".wallet__menu-filter").wait_for(
        state="visible",
        timeout=DEFAULT_TIMEOUT_MS,
    )

    page.locator(".recommendation-table").first.wait_for(
        state="visible",
        timeout=DEFAULT_TIMEOUT_MS,
    )


# -----------------------------------------------------------------------------
# Coleta das recommendation-table
# -----------------------------------------------------------------------------


def current_recommendation_table(page: Page):
    """Retorna a recommendation-table atualmente visível."""
    visible_table = page.locator(".recommendation-table:visible")

    if visible_table.count() > 0:
        return visible_table.first

    return page.locator(".recommendation-table").first


def visible_headers(page: Page) -> list[str]:
    """Obtém os títulos das colunas da recommendation-table visível."""
    table = current_recommendation_table(page)

    header_items = table.locator(".recommendation-table-header .header-item-wallet")

    headers: list[str] = []

    for index in range(header_items.count()):
        text = clean_text(header_items.nth(index).inner_text())

        if not text:
            continue

        # O tooltip pode entrar no inner_text.
        # A primeira linha é o nome efetivo da coluna.
        header = text.split("\n", 1)[0].strip()

        if header and header not in headers:
            headers.append(header)

    return headers


def visible_rows(page: Page) -> list[list[str]]:
    """
    Captura somente as linhas principais da tabela.

    O seletor direto evita:
    - coluna de botões;
    - conteúdo oculto do acordeão "Mais detalhes".
    """
    table = current_recommendation_table(page)

    row_locator = table.locator(
        ":scope > .recommendation-table-body > .body-contents > .body-contents-line"
    )

    rows: list[list[str]] = []

    for row_index in range(row_locator.count()):
        row = row_locator.nth(row_index)
        cells = row.locator(":scope > .value:not(.details)")

        values: list[str] = []

        for cell_index in range(cells.count()):
            values.append(clean_text(cells.nth(cell_index).inner_text()))

        if any(values):
            rows.append(values)

    return rows


def table_signature(rows: Iterable[Iterable[str]]) -> str:
    """Cria uma assinatura simples para detectar atualização da tabela."""
    return "||".join("|".join(row) for row in rows)


def active_category_text(page: Page) -> str:
    """Obtém o texto da categoria atualmente ativa."""
    active_tab = page.locator(".wallet__menu-filter .wallet__menu-filter-active")

    if active_tab.count() == 0:
        return ""

    return clean_text(active_tab.first.inner_text())


def click_category_tab(page: Page, category: str) -> None:
    """
    Ativa uma categoria usando click no DOM.

    O menu é um Swiper horizontal; algumas abas podem existir no DOM
    mas estar fora do viewport. element.click() no DOM evita o timeout.
    """
    candidates = page.locator(".wallet__menu-filter .swiper-slide").filter(
        has_text=category
    )

    if candidates.count() == 0:
        raise RuntimeError(f'Não encontrei a categoria "{category}".')

    candidates.first.evaluate("(element) => element.click()")


def wait_for_table_stable(
    page: Page,
    category: str,
) -> tuple[list[str], list[list[str]]]:
    """
    Aguarda a categoria ficar ativa e sua tabela estabilizar.

    Consideramos estável quando a mesma assinatura é observada em
    duas leituras consecutivas depois que a categoria ficou ativa.
    """
    deadline = time.monotonic() + TABLE_STABILIZE_TIMEOUT_SECONDS
    previous_signature: str | None = None
    stable_reads = 0

    latest_headers: list[str] = []
    latest_rows: list[list[str]] = []

    while time.monotonic() < deadline:
        page.wait_for_timeout(350)

        active_text = active_category_text(page)
        latest_headers = visible_headers(page)
        latest_rows = visible_rows(page)
        signature = table_signature(latest_rows)

        if active_text == category and latest_rows:
            if signature == previous_signature:
                stable_reads += 1
            else:
                stable_reads = 0
                previous_signature = signature

            if stable_reads >= 1:
                return latest_headers, latest_rows

    return latest_headers, latest_rows


def scrape_all_recommendation_tables(page: Page) -> list[RecommendationTable]:
    """
    Percorre todas as categorias disponíveis na área Ativos
    e coleta suas recommendation-table.
    """
    filter_selector = ".wallet__menu-filter .swiper-slide"
    filters = page.locator(filter_selector)

    categories: list[str] = []

    for index in range(filters.count()):
        category = clean_text(filters.nth(index).inner_text())

        if category and category not in categories:
            categories.append(category)

    if not categories:
        raise RuntimeError("Nenhuma categoria da carteira foi encontrada.")

    result: list[RecommendationTable] = []

    for category in categories:
        print(f"Coletando: {category}")

        click_category_tab(page, category)

        headers, rows = wait_for_table_stable(
            page,
            category,
        )

        max_columns = max(
            (len(row) for row in rows),
            default=0,
        )

        if len(headers) < max_columns:
            headers.extend(
                f"Coluna {column_number}"
                for column_number in range(
                    len(headers) + 1,
                    max_columns + 1,
                )
            )
        elif max_columns > 0 and len(headers) > max_columns:
            headers = headers[:max_columns]

        result.append(
            RecommendationTable(
                category=category,
                headers=headers,
                rows=rows,
            )
        )

    return result


# -----------------------------------------------------------------------------
# Excel da carteira atual
# -----------------------------------------------------------------------------


def style_sheet(
    worksheet: Worksheet,
    max_row: int,
    max_col: int,
) -> None:
    """Aplica visual semelhante à tabela escura da Finclass."""
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="202328",
    )

    body_fill = PatternFill(
        fill_type="solid",
        fgColor="0E1015",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
        size=11,
    )

    body_font = Font(
        color="FFFFFF",
        size=10,
    )

    thin = Side(
        style="thin",
        color="262A31",
    )

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    if max_row > 0 and max_col > 0:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=thin)

    worksheet.row_dimensions[1].height = 38

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
    ):
        for cell in row:
            cell.fill = body_fill
            cell.font = body_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = Border(bottom=thin)

    for row_index in range(2, max_row + 1):
        worksheet.row_dimensions[row_index].height = 42

    widths = {
        1: 24,
        2: 34,
        3: 25,
        4: 27,
        5: 24,
        6: 24,
    }

    for column_index in range(1, max_col + 1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = widths.get(
            column_index, 24
        )


def write_category_sheet(
    workbook: Workbook,
    table: RecommendationTable,
    used_names: set[str],
) -> None:
    """Cria uma aba para uma categoria específica."""
    category_name = table["category"]
    headers = table["headers"]
    rows = table["rows"]

    sheet_name = safe_sheet_name(
        category_name,
        used_names,
    )

    worksheet = workbook.create_sheet(sheet_name)
    effective_headers = headers if headers else ["Dados"]

    for column_index, header in enumerate(
        effective_headers,
        start=1,
    ):
        worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

    for row_index, values in enumerate(
        rows,
        start=2,
    ):
        for column_index, value in enumerate(
            values,
            start=1,
        ):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    max_col = max(
        len(effective_headers),
        max(
            (len(row) for row in rows),
            default=1,
        ),
    )

    max_row = max(
        len(rows) + 1,
        2,
    )

    style_sheet(
        worksheet,
        max_row=max_row,
        max_col=max_col,
    )


def write_consolidated_sheet(
    workbook: Workbook,
    tables: list[RecommendationTable],
) -> None:
    """Cria a aba Consolidado contendo todas as categorias."""
    worksheet = workbook.create_sheet(
        "Consolidado",
        0,
    )

    union_headers: list[str] = []

    for table in tables:
        for header in table["headers"]:
            if header not in union_headers:
                union_headers.append(header)

    consolidated_headers = [
        "Categoria",
        *union_headers,
    ]

    for column_index, header in enumerate(
        consolidated_headers,
        start=1,
    ):
        worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

    current_row = 2

    for table in tables:
        category = table["category"]
        table_headers = table["headers"]
        table_rows = table["rows"]

        for values in table_rows:
            row_map: dict[str, str] = {
                header: (values[index] if index < len(values) else "")
                for index, header in enumerate(table_headers)
            }

            worksheet.cell(
                row=current_row,
                column=1,
                value=category,
            )

            for column_index, header in enumerate(
                union_headers,
                start=2,
            ):
                worksheet.cell(
                    row=current_row,
                    column=column_index,
                    value=row_map.get(header, ""),
                )

            current_row += 1

    max_row = max(current_row - 1, 2)
    max_col = len(consolidated_headers)

    style_sheet(
        worksheet,
        max_row=max_row,
        max_col=max_col,
    )

    worksheet.column_dimensions["A"].width = 38


def export_excel(
    tables: list[RecommendationTable],
) -> Path:
    """Gera o arquivo Excel com a fotografia atual da carteira."""
    workbook = Workbook()

    # Workbook() cria uma worksheet inicial.
    # worksheets[0] evita o Optional de workbook.active no Pylance.
    workbook.remove(workbook.worksheets[0])

    used_names: set[str] = {"Consolidado"}

    write_consolidated_sheet(
        workbook,
        tables,
    )

    for table in tables:
        write_category_sheet(
            workbook,
            table,
            used_names,
        )

    workbook.properties.title = "Carteira Finclass"
    workbook.properties.subject = "Recomendações da Carteira Finclass"
    workbook.properties.creator = "Playwright + Python"
    workbook.properties.description = (
        "Dados atuais coletados automaticamente das "
        "recommendation-table da Carteira Finclass."
    )

    workbook.save(OUTPUT_FILE)

    return OUTPUT_FILE.resolve()


# -----------------------------------------------------------------------------
# Histórico diário e detecção de alterações
# -----------------------------------------------------------------------------


def parse_percentage(text: str) -> float | None:
    """
    Converte textos como:
      1,00%\nR$ 2.500
      9.47%

    para decimal do Excel:
      0.01
      0.0947
    """
    match = re.search(r"(-?\d+(?:[.,]\d+)?)\s*%", text)

    if not match:
        return None

    normalized = match.group(1).replace(",", ".")

    try:
        return float(normalized) / 100.0
    except ValueError:
        return None


def parse_company_and_ticker(text: str) -> tuple[str, str]:
    """
    Converte:
      BR PARTNERS\nBRBI11

    para:
      ("BR PARTNERS", "BRBI11")
    """
    normalized = clean_text(text)
    lines = normalized.splitlines()

    if not lines:
        return "", ""

    company = lines[0].strip()
    ticker = lines[1].strip() if len(lines) > 1 else ""

    return company, ticker


def build_asset_records(
    tables: list[RecommendationTable],
) -> list[AssetRecord]:
    """
    Extrai todos os nomes de ativos/fundos/empresas da carteira atual.

    Preserva todas as linhas coletadas. A aba "Nomes únicos" é deduplicada
    separadamente no momento da exportação.
    """
    records: list[AssetRecord] = []

    for table in tables:
        category = table["category"]

        for row_number, row in enumerate(table["rows"], start=1):
            if len(row) < 2:
                raise RuntimeError(
                    "Não foi possível extrair o nome do ativo em "
                    f"{category} | linha {row_number} | {row!r}"
                )

            name, code = parse_company_and_ticker(row[1])

            if not name:
                raise RuntimeError(
                    f"Nome vazio em {category} | linha {row_number} | {row[1]!r}"
                )

            records.append(
                AssetRecord(
                    category=category,
                    name=name,
                    code=code,
                )
            )

    return records


def unique_asset_records(
    records: list[AssetRecord],
) -> list[AssetRecord]:
    """
    Retorna nomes únicos sem perder itens diferentes que compartilhem
    o mesmo nome.

    A chave considera categoria + nome + código.
    """
    unique: list[AssetRecord] = []
    seen: set[str] = set()

    for record in records:
        key = "{}|{}|{}".format(
            record["category"].strip().casefold(),
            record["name"].strip().casefold(),
            record["code"].strip().upper(),
        )

        if key in seen:
            continue

        seen.add(key)
        unique.append(record)

    return unique


def export_asset_catalog(
    tables: list[RecommendationTable],
) -> tuple[Path, int, int]:
    """
    Gera finclass_ativos.txt.

    O arquivo contém os nomes atuais de ativos/fundos/empresas,
    organizados por categoria, no formato:

        NOME | CÓDIGO

    A lista escrita é deduplicada por:
        categoria + nome + código

    O total bruto de itens processados também é retornado para permitir
    validar que todos os itens coletados foram analisados.
    """
    all_records = build_asset_records(tables)
    unique_records = unique_asset_records(all_records)

    by_category: dict[str, list[AssetRecord]] = {}

    for record in unique_records:
        category = record["category"]

        if category not in by_category:
            by_category[category] = []

        by_category[category].append(record)

    lines: list[str] = [
        "FINCLASS - ATIVOS / FUNDOS / EMPRESAS",
        "Gerado em: {}".format(datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        f"Itens coletados: {len(all_records)}",
        f"Nomes únicos: {len(unique_records)}",
        "",
    ]

    for category, records in by_category.items():
        lines.append(f"[{category}]")

        for record in sorted(
            records,
            key=lambda item: (
                item["name"].casefold(),
                item["code"].upper(),
            ),
        ):
            name = record["name"]
            code = record["code"]

            if code:
                lines.append(f"{name} | {code}")
            else:
                lines.append(name)

        lines.append("")

    ASSETS_FILE.write_text(
        "\n".join(lines).rstrip() + "\n",
        encoding="utf-8",
    )

    return (
        ASSETS_FILE.resolve(),
        len(all_records),
        len(unique_records),
    )


def history_key(record: HistoryRecord) -> str:
    """Chave estável para comparar um ativo entre dois dias."""
    category = record["category"].strip().casefold()
    ticker = record["ticker"].strip().upper()
    company = record["company"].strip().casefold()

    identifier = ticker if ticker else company

    return f"{category}|{identifier}"


def build_daily_snapshot(
    tables: list[RecommendationTable],
    snapshot_date: date,
) -> list[HistoryRecord]:
    """
    Monta o snapshot diário preservando TODAS as linhas coletadas.

    Regras importantes:
    - não remove duplicidades;
    - não substitui registros com o mesmo ticker/nome;
    - não ignora silenciosamente linhas que não possam ser convertidas;
    - se o scraper coletou N linhas, o snapshot deve conter exatamente N linhas.

    Campos gravados:
      Data + Categoria + % da Carteira + Nome da empresa + Código
    """
    records: list[HistoryRecord] = []
    invalid_rows: list[str] = []
    total_collected = 0

    for table in tables:
        category = table["category"]

        for row_number, row in enumerate(table["rows"], start=1):
            total_collected += 1

            # Estrutura observada na recommendation-table:
            # 0 = % da Carteira + valor em R$
            # 1 = Nome da empresa + ticker/código
            if len(row) < 2:
                invalid_rows.append(
                    f"{category} | linha {row_number} | menos de 2 colunas | {row!r}"
                )
                continue

            percentage = parse_percentage(row[0])
            company, ticker = parse_company_and_ticker(row[1])

            if percentage is None:
                invalid_rows.append(
                    f"{category} | linha {row_number} | percentual inválido | {row[0]!r}"
                )
                continue

            if not company:
                invalid_rows.append(
                    f"{category} | linha {row_number} | nome da empresa vazio | {row[1]!r}"
                )
                continue

            records.append(
                HistoryRecord(
                    snapshot_date=snapshot_date,
                    category=category,
                    percentage=percentage,
                    company=company,
                    ticker=ticker,
                )
            )

    # Nunca grava um histórico parcial sem avisar. Isso garante que o número
    # de itens do snapshot seja exatamente o número de itens coletados.
    if invalid_rows or len(records) != total_collected:
        details = "\n".join(f"  - {item}" for item in invalid_rows[:20])

        if len(invalid_rows) > 20:
            details += f"\n  - ... e mais {len(invalid_rows) - 20} linha(s) inválida(s)"

        raise RuntimeError(
            "Snapshot histórico incompleto: o scraper coletou {} item(ns), "
            "mas apenas {} puderam ser convertidos para o histórico.\n{}".format(
                total_collected,
                len(records),
                details or "  - motivo não identificado",
            )
        )

    return records


def normalize_excel_date(value: object) -> date | None:
    """Normaliza uma data lida do Excel para datetime.date."""
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, str):
        for date_format in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(value.strip(), date_format).date()
            except ValueError:
                continue

    return None


def normalize_excel_percentage(value: object) -> float | None:
    """Normaliza percentual lido de uma planilha antiga/nova."""
    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        return parse_percentage(value)

    return None


def get_or_create_worksheet(
    workbook: Workbook,
    sheet_name: str,
) -> Worksheet:
    """Retorna uma Worksheet existente ou cria uma nova."""
    if sheet_name in workbook.sheetnames:
        return cast(Worksheet, workbook[sheet_name])

    return workbook.create_sheet(sheet_name)


def read_history_records(worksheet: Worksheet) -> list[HistoryRecord]:
    """Lê os snapshots existentes da aba Histórico."""
    records: list[HistoryRecord] = []

    if worksheet.max_row < 2:
        return records

    for row_index in range(2, worksheet.max_row + 1):
        snapshot_date = normalize_excel_date(
            worksheet.cell(row=row_index, column=1).value
        )
        category_value = worksheet.cell(row=row_index, column=2).value
        percentage = normalize_excel_percentage(
            worksheet.cell(row=row_index, column=3).value
        )
        company_value = worksheet.cell(row=row_index, column=4).value
        ticker_value = worksheet.cell(row=row_index, column=5).value

        if snapshot_date is None or percentage is None or not company_value:
            continue

        records.append(
            HistoryRecord(
                snapshot_date=snapshot_date,
                category=str(category_value or ""),
                percentage=percentage,
                company=str(company_value),
                ticker=str(ticker_value or ""),
            )
        )

    return records


def previous_snapshot(
    history_records: list[HistoryRecord],
    current_date: date,
) -> tuple[date | None, list[HistoryRecord]]:
    """Obtém o último snapshot anterior ao dia atual."""
    previous_dates = {
        record["snapshot_date"]
        for record in history_records
        if record["snapshot_date"] < current_date
    }

    if not previous_dates:
        return None, []

    previous_date = max(previous_dates)
    records = [
        record for record in history_records if record["snapshot_date"] == previous_date
    ]

    return previous_date, records


def _same_percentage(left: float, right: float) -> bool:
    """Compara percentuais com tolerância para ponto flutuante."""
    return abs(left - right) <= 0.0000001


def _group_snapshot_records(
    records: list[HistoryRecord],
) -> dict[str, list[HistoryRecord]]:
    """
    Agrupa registros sem perder duplicidades.

    Diferentemente de um dict[str, HistoryRecord], cada chave guarda uma lista.
    Assim, dois ou mais itens com o mesmo ticker/nome continuam existindo.
    """
    grouped: dict[str, list[HistoryRecord]] = {}

    for record in records:
        key = history_key(record)
        grouped.setdefault(key, []).append(record)

    return grouped


def _append_change(
    changes: list[ChangeRecord],
    previous_date: date,
    current_date: date,
    record: HistoryRecord,
    change_type: str,
    previous_percentage: float | None,
    current_percentage: float | None,
) -> None:
    changes.append(
        ChangeRecord(
            change_date=current_date,
            previous_date=previous_date,
            category=record["category"],
            company=record["company"],
            ticker=record["ticker"],
            change_type=change_type,
            previous_percentage=previous_percentage,
            current_percentage=current_percentage,
        )
    )


def compare_snapshots(
    old_records: list[HistoryRecord],
    new_records: list[HistoryRecord],
    previous_date: date,
    current_date: date,
) -> list[ChangeRecord]:
    """
    Compara dois snapshots SEM colapsar itens repetidos.

    Detecta:
      - ADICIONADO
      - REMOVIDO
      - % ALTERADA
      - NOME ALTERADO (quando há ticker estável)

    Estratégia para chaves repetidas:
    1. pares idênticos de percentual são consumidos primeiro;
    2. registros restantes são pareados para detectar alteração de percentual;
    3. sobras de um lado são adições/remoções.

    Dessa forma, a comparação funciona como um multiconjunto e não perde
    registros quando o mesmo ticker/nome aparece mais de uma vez.
    """
    old_groups = _group_snapshot_records(old_records)
    new_groups = _group_snapshot_records(new_records)

    changes: list[ChangeRecord] = []
    all_keys = sorted(set(old_groups) | set(new_groups))

    for key in all_keys:
        old_group = list(old_groups.get(key, []))
        new_group = list(new_groups.get(key, []))

        # Primeiro remove pares exatamente equivalentes.
        unmatched_old: list[HistoryRecord] = []
        available_new = list(new_group)

        for old_record in old_group:
            exact_index: int | None = None

            for index, new_record in enumerate(available_new):
                same_percentage = _same_percentage(
                    old_record["percentage"],
                    new_record["percentage"],
                )
                same_name = (
                    old_record["company"].strip().casefold()
                    == new_record["company"].strip().casefold()
                )

                if same_percentage and same_name:
                    exact_index = index
                    break

            if exact_index is None:
                unmatched_old.append(old_record)
            else:
                available_new.pop(exact_index)

        unmatched_new = available_new

        # Ordenação determinística para parear múltiplas ocorrências da mesma chave.
        unmatched_old.sort(
            key=lambda item: (
                item["percentage"],
                item["company"].casefold(),
                item["ticker"],
            )
        )
        unmatched_new.sort(
            key=lambda item: (
                item["percentage"],
                item["company"].casefold(),
                item["ticker"],
            )
        )

        pair_count = min(len(unmatched_old), len(unmatched_new))

        for index in range(pair_count):
            old_record = unmatched_old[index]
            new_record = unmatched_new[index]

            percentage_changed = not _same_percentage(
                old_record["percentage"],
                new_record["percentage"],
            )
            name_changed = (
                old_record["company"].strip().casefold()
                != new_record["company"].strip().casefold()
            )

            if percentage_changed:
                _append_change(
                    changes,
                    previous_date,
                    current_date,
                    new_record,
                    "% ALTERADA",
                    old_record["percentage"],
                    new_record["percentage"],
                )

            if name_changed:
                _append_change(
                    changes,
                    previous_date,
                    current_date,
                    new_record,
                    "NOME ALTERADO",
                    old_record["percentage"],
                    new_record["percentage"],
                )

        # Registros novos que não encontraram par.
        for new_record in unmatched_new[pair_count:]:
            _append_change(
                changes,
                previous_date,
                current_date,
                new_record,
                "ADICIONADO",
                None,
                new_record["percentage"],
            )

        # Registros antigos que não encontraram par.
        for old_record in unmatched_old[pair_count:]:
            _append_change(
                changes,
                previous_date,
                current_date,
                old_record,
                "REMOVIDO",
                old_record["percentage"],
                None,
            )

    return changes


def delete_rows_for_date(
    worksheet: Worksheet,
    target_date: date,
    date_column: int,
) -> None:
    """
    Remove registros de uma data específica.

    Isso torna a execução idempotente: se o script rodar 3 vezes no mesmo
    dia, o histórico terá apenas um snapshot desse dia.
    """
    for row_index in range(worksheet.max_row, 1, -1):
        row_date = normalize_excel_date(
            worksheet.cell(
                row=row_index,
                column=date_column,
            ).value
        )

        if row_date == target_date:
            worksheet.delete_rows(row_index, 1)


def ensure_history_header(worksheet: Worksheet) -> None:
    headers = [
        "Data",
        "Categoria",
        "% da Carteira",
        "Nome da empresa",
        "Código",
    ]

    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )


def ensure_changes_header(worksheet: Worksheet) -> None:
    headers = [
        "Data da alteração",
        "Comparado com",
        "Categoria",
        "Nome da empresa",
        "Código",
        "Alteração",
        "% anterior",
        "% atual",
    ]

    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )


def style_history_worksheet(worksheet: Worksheet) -> None:
    """Formata a aba Histórico."""
    ensure_history_header(worksheet)

    max_row = max(worksheet.max_row, 1)
    max_col = 5

    style_sheet(
        worksheet,
        max_row=max_row,
        max_col=max_col,
    )

    worksheet.column_dimensions["A"].width = 14
    worksheet.column_dimensions["B"].width = 40
    worksheet.column_dimensions["C"].width = 18
    worksheet.column_dimensions["D"].width = 34
    worksheet.column_dimensions["E"].width = 16

    for row_index in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row_index, column=1).number_format = "dd/mm/yyyy"
        worksheet.cell(row=row_index, column=3).number_format = "0.00%"


def style_changes_worksheet(worksheet: Worksheet) -> None:
    """Formata a aba Alterações."""
    ensure_changes_header(worksheet)

    max_row = max(worksheet.max_row, 1)
    max_col = 8

    style_sheet(
        worksheet,
        max_row=max_row,
        max_col=max_col,
    )

    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 18
    worksheet.column_dimensions["C"].width = 40
    worksheet.column_dimensions["D"].width = 34
    worksheet.column_dimensions["E"].width = 16
    worksheet.column_dimensions["F"].width = 20
    worksheet.column_dimensions["G"].width = 16
    worksheet.column_dimensions["H"].width = 16

    for row_index in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row_index, column=1).number_format = "dd/mm/yyyy"
        worksheet.cell(row=row_index, column=2).number_format = "dd/mm/yyyy"
        worksheet.cell(row=row_index, column=7).number_format = "0.00%"
        worksheet.cell(row=row_index, column=8).number_format = "0.00%"


def save_daily_history(
    tables: list[RecommendationTable],
    current_date: date,
) -> tuple[Path, date | None, list[ChangeRecord], int]:
    """
    Salva o snapshot do dia e registra somente as mudanças detectadas.

    Regras:
    - um único snapshot por dia;
    - se executar novamente no mesmo dia, substitui o snapshot daquele dia;
    - compara sempre com a última data anterior disponível;
    - o primeiro dia vira a base inicial, sem gerar falsos "ADICIONADO".
    """
    current_snapshot = build_daily_snapshot(
        tables,
        current_date,
    )

    if not current_snapshot:
        raise RuntimeError(
            "Não foi possível montar o histórico: nenhuma linha válida "
            "com % da Carteira e Nome da empresa foi encontrada."
        )

    if HISTORY_FILE.exists():
        workbook = load_workbook(HISTORY_FILE)
    else:
        workbook = Workbook()
        initial_worksheet = workbook.worksheets[0]
        initial_worksheet.title = HISTORY_SHEET

    history_worksheet = get_or_create_worksheet(
        workbook,
        HISTORY_SHEET,
    )
    changes_worksheet = get_or_create_worksheet(
        workbook,
        CHANGES_SHEET,
    )

    ensure_history_header(history_worksheet)
    ensure_changes_header(changes_worksheet)

    existing_history = read_history_records(history_worksheet)

    previous_date, old_snapshot = previous_snapshot(
        existing_history,
        current_date,
    )

    changes: list[ChangeRecord] = []

    if previous_date is not None:
        changes = compare_snapshots(
            old_snapshot,
            current_snapshot,
            previous_date,
            current_date,
        )

    # Idempotência: substitui o snapshot e as alterações do dia atual.
    delete_rows_for_date(
        history_worksheet,
        current_date,
        date_column=1,
    )
    delete_rows_for_date(
        changes_worksheet,
        current_date,
        date_column=1,
    )

    for record in current_snapshot:
        history_worksheet.append(
            [
                record["snapshot_date"],
                record["category"],
                record["percentage"],
                record["company"],
                record["ticker"],
            ]
        )

    for change in changes:
        changes_worksheet.append(
            [
                change["change_date"],
                change["previous_date"],
                change["category"],
                change["company"],
                change["ticker"],
                change["change_type"],
                change["previous_percentage"],
                change["current_percentage"],
            ]
        )

    style_history_worksheet(history_worksheet)
    style_changes_worksheet(changes_worksheet)

    workbook.properties.title = "Histórico da Carteira Finclass"
    workbook.properties.subject = "Snapshots diários e alterações da carteira"
    workbook.properties.creator = "Playwright + Python"
    workbook.properties.description = (
        "Histórico diário da porcentagem recomendada por ativo, "
        "com detecção de adições, remoções e alterações de percentual."
    )

    workbook.save(HISTORY_FILE)

    return (
        HISTORY_FILE.resolve(),
        previous_date,
        changes,
        len(current_snapshot),
    )


# -----------------------------------------------------------------------------
# Notificação por e-mail
# -----------------------------------------------------------------------------


def get_email_settings() -> EmailSettings:
    """Carrega e valida a configuração SMTP do .env."""
    host = required_env("SMTP_HOST")
    username = required_env("SMTP_USER")
    password = required_env("SMTP_PASSWORD")
    sender = os.getenv("EMAIL_FROM", username).strip()
    sender_name = os.getenv(
        "EMAIL_FROM_NAME",
        "Finclass Monitor",
    ).strip()

    recipients_text = required_env("EMAIL_TO")
    recipients = [item.strip() for item in recipients_text.split(",") if item.strip()]

    if not recipients:
        raise RuntimeError("EMAIL_TO não possui nenhum destinatário válido.")

    port_text = os.getenv("SMTP_PORT", "587").strip()

    try:
        port = int(port_text)
    except ValueError as exc:
        raise RuntimeError("SMTP_PORT deve ser um número inteiro.") from exc

    security = (
        os.getenv(
            "SMTP_SECURITY",
            "starttls",
        )
        .strip()
        .lower()
    )

    if security not in {"starttls", "ssl", "none"}:
        raise RuntimeError("SMTP_SECURITY deve ser starttls, ssl ou none.")

    return EmailSettings(
        host=host,
        port=port,
        security=security,
        username=username,
        password=password,
        sender=sender,
        sender_name=sender_name,
        recipients=recipients,
    )


def percentage_text(value: float | None) -> str:
    """Formata um percentual para exibição no e-mail/terminal."""
    if value is None:
        return "—"

    return f"{value * 100:.2f}%"


def change_sort_key(change: ChangeRecord) -> tuple[str, str, str]:
    """Ordenação consistente das alterações no e-mail."""
    return (
        change["change_type"],
        change["category"].casefold(),
        (change["ticker"] or change["company"]).casefold(),
    )


def changes_signature(
    changes: list[ChangeRecord],
    previous_date: date,
    current_date: date,
) -> str:
    """
    Gera uma assinatura determinística.

    Evita enviar o mesmo e-mail várias vezes caso o script seja executado
    repetidamente no mesmo dia sem novas alterações.
    """
    payload = {
        "previous_date": previous_date.isoformat(),
        "current_date": current_date.isoformat(),
        "changes": [
            {
                "category": item["category"],
                "company": item["company"],
                "ticker": item["ticker"],
                "change_type": item["change_type"],
                "previous_percentage": item["previous_percentage"],
                "current_percentage": item["current_percentage"],
            }
            for item in sorted(changes, key=change_sort_key)
        ],
    }

    serialized = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )

    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


def email_was_already_sent(
    signature: str,
    current_date: date,
) -> bool:
    """Verifica se a mesma alteração já foi notificada."""
    if not EMAIL_STATE_FILE.exists():
        return False

    try:
        payload = json.loads(EMAIL_STATE_FILE.read_text(encoding="utf-8"))
    except (OSError, ValueError, TypeError):
        return False

    return (
        payload.get("date") == current_date.isoformat()
        and payload.get("signature") == signature
    )


def mark_email_as_sent(
    signature: str,
    current_date: date,
) -> None:
    """Persiste a assinatura somente depois que o SMTP confirmou o envio."""
    payload = {
        "date": current_date.isoformat(),
        "signature": signature,
        "sent_at": datetime.now().isoformat(timespec="seconds"),
    }

    EMAIL_STATE_FILE.write_text(
        json.dumps(
            payload,
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def build_email_subject(
    changes: list[ChangeRecord],
    current_date: date,
) -> str:
    """Assunto curto e informativo."""
    return ("[Finclass] Carteira alterada — {} — {} alteração(ões)").format(
        current_date.strftime("%d/%m/%Y"),
        len(changes),
    )


def build_email_text(
    changes: list[ChangeRecord],
    previous_date: date,
    current_date: date,
) -> str:
    """Versão texto puro para clientes que não renderizam HTML."""
    lines = [
        "Carteira Finclass alterada",
        "",
        "Comparação: {} x {}".format(
            previous_date.strftime("%d/%m/%Y"),
            current_date.strftime("%d/%m/%Y"),
        ),
        f"Alterações detectadas: {len(changes)}",
        "",
    ]

    for change in sorted(changes, key=change_sort_key):
        lines.append(
            "{} | {} | {} | {} -> {}".format(
                change["change_type"],
                change["ticker"] or "sem código",
                change["company"],
                percentage_text(change["previous_percentage"]),
                percentage_text(change["current_percentage"]),
            )
        )

    lines.extend(
        [
            "",
            "E-mail gerado automaticamente pelo Finclass Monitor.",
        ]
    )

    return "\n".join(lines)


def build_email_html(
    changes: list[ChangeRecord],
    previous_date: date,
    current_date: date,
) -> str:
    """Gera o e-mail HTML formatado com todas as alterações."""
    counts: dict[str, int] = {}

    for change in changes:
        change_type = change["change_type"]
        counts[change_type] = counts.get(change_type, 0) + 1

    badge_order = [
        "ADICIONADO",
        "REMOVIDO",
        "% ALTERADA",
        "NOME ALTERADO",
    ]

    badges = []

    for change_type in badge_order:
        count = counts.get(change_type, 0)

        if count == 0:
            continue

        badges.append(
            f"""
            <div style="
                display:inline-block;
                margin:0 8px 8px 0;
                padding:10px 14px;
                background:#f3f4f6;
                border:1px solid #e5e7eb;
                border-radius:8px;
                font-size:13px;
                color:#111827;
            ">
                <strong>{escape(change_type)}</strong>: {count}
            </div>
            """
        )

    rows = []

    for change in sorted(changes, key=change_sort_key):
        rows.append(
            """
            <tr>
                <td style="{cell}">{change_type}</td>
                <td style="{cell}">{category}</td>
                <td style="{cell}">
                    <strong>{company}</strong>
                    {ticker}
                </td>
                <td style="{cell}; text-align:center;">{previous}</td>
                <td style="{cell}; text-align:center;"><strong>{current}</strong></td>
            </tr>
            """.format(
                cell=(
                    "padding:12px 10px;"
                    "border-bottom:1px solid #e5e7eb;"
                    "font-size:13px;"
                    "color:#111827"
                ),
                change_type=escape(change["change_type"]),
                category=escape(change["category"]),
                company=escape(change["company"]),
                ticker=(
                    "<br><span style='color:#6b7280'>{}</span>".format(
                        escape(change["ticker"])
                    )
                    if change["ticker"]
                    else ""
                ),
                previous=escape(percentage_text(change["previous_percentage"])),
                current=escape(percentage_text(change["current_percentage"])),
            )
        )

    return """
    <!doctype html>
    <html lang="pt-BR">
      <body style="
          margin:0;
          padding:0;
          background:#f5f7fa;
          font-family:Arial,Helvetica,sans-serif;
          color:#111827;
      ">
        <div style="padding:28px 12px;">
          <div style="
              max-width:920px;
              margin:0 auto;
              background:#ffffff;
              border:1px solid #e5e7eb;
              border-radius:12px;
              overflow:hidden;
          ">
            <div style="
                background:#111827;
                color:#ffffff;
                padding:24px 28px;
            ">
              <div style="
                  font-size:12px;
                  text-transform:uppercase;
                  letter-spacing:1px;
                  opacity:.75;
              ">
                Finclass Monitor
              </div>
              <h1 style="
                  margin:8px 0 4px;
                  font-size:24px;
                  line-height:1.25;
              ">
                Alteração detectada na carteira
              </h1>
              <div style="
                  font-size:14px;
                  opacity:.85;
              ">
                Comparação: {previous_date} → {current_date}
              </div>
            </div>

            <div style="padding:24px 28px;">
              <p style="
                  margin:0 0 18px;
                  font-size:15px;
                  line-height:1.6;
              ">
                Foram detectadas <strong>{change_count}</strong>
                alteração(ões) em relação ao último snapshot disponível.
              </p>

              <div style="margin:0 0 18px;">
                {badges}
              </div>

              <div style="
                  overflow-x:auto;
                  border:1px solid #e5e7eb;
                  border-radius:10px;
              ">
                <table style="
                    width:100%;
                    border-collapse:collapse;
                    min-width:760px;
                ">
                  <thead>
                    <tr style="background:#f9fafb;">
                      <th style="{header}">Alteração</th>
                      <th style="{header}">Categoria</th>
                      <th style="{header}">Ativo</th>
                      <th style="{header}; text-align:center;">% anterior</th>
                      <th style="{header}; text-align:center;">% atual</th>
                    </tr>
                  </thead>
                  <tbody>
                    {rows}
                  </tbody>
                </table>
              </div>

              <p style="
                  margin:22px 0 0;
                  color:#6b7280;
                  font-size:12px;
                  line-height:1.5;
              ">
                O histórico completo permanece salvo em
                <strong>finclass_historico.xlsx</strong>.
                Este e-mail foi gerado automaticamente.
              </p>
            </div>
          </div>
        </div>
      </body>
    </html>
    """.format(
        previous_date=previous_date.strftime("%d/%m/%Y"),
        current_date=current_date.strftime("%d/%m/%Y"),
        change_count=len(changes),
        badges="".join(badges),
        rows="".join(rows),
        header=(
            "padding:11px 10px;"
            "border-bottom:1px solid #e5e7eb;"
            "font-size:12px;"
            "text-align:left;"
            "color:#374151"
        ),
    )


def smtp_send_message(
    message: EmailMessage,
    settings: EmailSettings,
) -> None:
    """Envia uma mensagem usando as configurações SMTP validadas."""
    security = settings["security"]

    if security == "ssl":
        with smtplib.SMTP_SSL(
            settings["host"],
            settings["port"],
            timeout=30,
            context=ssl.create_default_context(),
        ) as smtp:
            smtp.login(
                settings["username"],
                settings["password"],
            )
            smtp.send_message(message)
        return

    with smtplib.SMTP(
        settings["host"],
        settings["port"],
        timeout=30,
    ) as smtp:
        smtp.ehlo()

        if security == "starttls":
            smtp.starttls(context=ssl.create_default_context())
            smtp.ehlo()

        smtp.login(
            settings["username"],
            settings["password"],
        )
        smtp.send_message(message)


def send_test_email() -> None:
    """
    Testa somente o SMTP.

    Não acessa a Finclass.
    Não altera finclass_historico.xlsx.
    Não altera finclass_email_state.json.
    """
    if not ENV_FILE.exists():
        raise RuntimeError(f"Arquivo .env não encontrado em: {ENV_FILE}")

    settings = get_email_settings()
    now = datetime.now()

    message = EmailMessage()
    message["Subject"] = "[Finclass] Teste de e-mail — {}".format(
        now.strftime("%d/%m/%Y %H:%M:%S")
    )
    message["From"] = "{} <{}>".format(
        settings["sender_name"],
        settings["sender"],
    )
    message["To"] = ", ".join(settings["recipients"])

    message.set_content(
        "\n".join(
            [
                "Finclass Monitor",
                "",
                "Teste de envio concluído.",
                "Se você recebeu esta mensagem, a configuração SMTP está funcionando.",
                "",
                "Servidor: {}:{}".format(
                    settings["host"],
                    settings["port"],
                ),
                "Segurança: {}".format(settings["security"]),
                "Data/hora: {}".format(now.strftime("%d/%m/%Y %H:%M:%S")),
            ]
        )
    )

    message.add_alternative(
        """
        <!doctype html>
        <html lang="pt-BR">
          <body style="
              margin:0;
              padding:0;
              background:#f5f7fa;
              font-family:Arial,Helvetica,sans-serif;
              color:#111827;
          ">
            <div style="padding:32px 12px;">
              <div style="
                  max-width:640px;
                  margin:0 auto;
                  background:#ffffff;
                  border:1px solid #e5e7eb;
                  border-radius:12px;
                  overflow:hidden;
              ">
                <div style="
                    background:#111827;
                    color:#ffffff;
                    padding:24px 28px;
                ">
                  <div style="
                      font-size:12px;
                      text-transform:uppercase;
                      letter-spacing:1px;
                      opacity:.75;
                  ">
                    Finclass Monitor
                  </div>
                  <h1 style="
                      margin:8px 0 0;
                      font-size:24px;
                  ">
                    Teste de e-mail concluído
                  </h1>
                </div>

                <div style="padding:26px 28px;">
                  <p style="
                      margin:0 0 16px;
                      font-size:15px;
                      line-height:1.6;
                  ">
                    Se você recebeu esta mensagem, a configuração
                    <strong>SMTP do Gmail está funcionando</strong>.
                  </p>

                  <table style="
                      width:100%;
                      border-collapse:collapse;
                      font-size:14px;
                  ">
                    <tr>
                      <td style="padding:8px 0;color:#6b7280;">Servidor</td>
                      <td style="padding:8px 0;text-align:right;">
                        {host}:{port}
                      </td>
                    </tr>
                    <tr>
                      <td style="padding:8px 0;color:#6b7280;">Segurança</td>
                      <td style="padding:8px 0;text-align:right;">
                        {security}
                      </td>
                    </tr>
                    <tr>
                      <td style="padding:8px 0;color:#6b7280;">Data/hora</td>
                      <td style="padding:8px 0;text-align:right;">
                        {datetime}
                      </td>
                    </tr>
                  </table>
                </div>
              </div>
            </div>
          </body>
        </html>
        """.format(
            host=escape(settings["host"]),
            port=settings["port"],
            security=escape(settings["security"]),
            datetime=escape(now.strftime("%d/%m/%Y %H:%M:%S")),
        ),
        subtype="html",
    )

    print(
        "Testando SMTP {}:{} ({})...".format(
            settings["host"],
            settings["port"],
            settings["security"],
        )
    )

    try:
        smtp_send_message(
            message,
            settings,
        )
    except smtplib.SMTPAuthenticationError as exc:
        raise RuntimeError(
            "Falha de autenticação SMTP do Gmail. "
            "Confira SMTP_USER e use uma Senha de app válida do Google."
        ) from exc
    except (smtplib.SMTPException, OSError) as exc:
        raise RuntimeError(f"Falha no envio SMTP: {exc}") from exc

    print(
        "E-mail de teste enviado com sucesso para: {}".format(
            ", ".join(settings["recipients"])
        )
    )


def send_simulated_change_email() -> None:
    """
    Envia um e-mail usando exatamente o layout de alterações da produção.

    Regras do modo de simulação:
    - não acessa a Finclass;
    - não altera finclass_carteira.xlsx;
    - não altera finclass_historico.xlsx;
    - não altera finclass_email_state.json;
    - não interfere na prevenção de e-mails duplicados da produção.
    """
    if not ENV_FILE.exists():
        raise RuntimeError(f"Arquivo .env não encontrado em: {ENV_FILE}")

    settings = get_email_settings()

    current_date = date.today()
    previous_date = current_date - timedelta(days=1)

    changes: list[ChangeRecord] = [
        ChangeRecord(
            change_date=current_date,
            previous_date=previous_date,
            category="A - AÇÕES",
            company="BR PARTNERS",
            ticker="BRBI11",
            change_type="% ALTERADA",
            previous_percentage=0.0100,
            current_percentage=0.0150,
        ),
        ChangeRecord(
            change_date=current_date,
            previous_date=previous_date,
            category="A - AÇÕES",
            company="EMPRESA EXEMPLO",
            ticker="TEST3",
            change_type="ADICIONADO",
            previous_percentage=None,
            current_percentage=0.0200,
        ),
        ChangeRecord(
            change_date=current_date,
            previous_date=previous_date,
            category="R - REAL ESTATE (FUNDOS IMOBILIÁRIOS)",
            company="FUNDO EXEMPLO",
            ticker="TEST11",
            change_type="REMOVIDO",
            previous_percentage=0.0300,
            current_percentage=None,
        ),
    ]

    message = EmailMessage()

    message["Subject"] = (
        "[Finclass] SIMULAÇÃO — Carteira alterada — {} — {} alteração(ões)"
    ).format(
        current_date.strftime("%d/%m/%Y"),
        len(changes),
    )

    message["From"] = "{} <{}>".format(
        settings["sender_name"],
        settings["sender"],
    )

    message["To"] = ", ".join(settings["recipients"])

    message.set_content(
        build_email_text(
            changes,
            previous_date,
            current_date,
        )
        + "\n\n[SIMULAÇÃO] Nenhum histórico foi modificado."
    )

    simulated_html = build_email_html(
        changes,
        previous_date,
        current_date,
    )

    # Identifica visualmente o e-mail como simulação sem alterar
    # o layout usado pela notificação real.
    simulated_html = simulated_html.replace(
        "Alteração detectada na carteira",
        "SIMULAÇÃO — Alteração detectada na carteira",
        1,
    ).replace(
        "Este e-mail foi gerado automaticamente.",
        (
            "Este é um e-mail de <strong>SIMULAÇÃO</strong>. "
            "Nenhum arquivo de histórico foi modificado."
        ),
        1,
    )

    message.add_alternative(
        simulated_html,
        subtype="html",
    )

    print(
        "Enviando simulação de alteração para: {}".format(
            ", ".join(settings["recipients"])
        )
    )

    try:
        smtp_send_message(
            message,
            settings,
        )
    except smtplib.SMTPAuthenticationError as exc:
        raise RuntimeError(
            "Falha de autenticação SMTP do Gmail. Confira SMTP_USER e a Senha de app."
        ) from exc
    except (smtplib.SMTPException, OSError) as exc:
        raise RuntimeError(f"Falha no envio SMTP: {exc}") from exc

    print("E-mail de simulação enviado com sucesso.")
    print("Nenhum arquivo Excel ou estado de notificação foi alterado.")


def send_changes_email(
    changes: list[ChangeRecord],
    previous_date: date,
    current_date: date,
) -> bool:
    """
    Envia e-mail somente quando:
    - EMAIL_ENABLED=true;
    - existe snapshot anterior;
    - existem alterações;
    - a mesma assinatura ainda não foi enviada.

    Retorna True somente quando um novo e-mail foi enviado.
    """
    if not env_flag("EMAIL_ENABLED", True):
        print("E-mail: envio desabilitado por EMAIL_ENABLED=false.")
        return False

    if not changes:
        return False

    signature = changes_signature(
        changes,
        previous_date,
        current_date,
    )

    if email_was_already_sent(signature, current_date):
        print(
            "E-mail: estas mesmas alterações já foram notificadas hoje; "
            "nenhum e-mail duplicado foi enviado."
        )
        return False

    settings = get_email_settings()

    message = EmailMessage()
    message["Subject"] = build_email_subject(
        changes,
        current_date,
    )
    message["From"] = "{} <{}>".format(
        settings["sender_name"],
        settings["sender"],
    )
    message["To"] = ", ".join(settings["recipients"])

    message.set_content(
        build_email_text(
            changes,
            previous_date,
            current_date,
        )
    )

    message.add_alternative(
        build_email_html(
            changes,
            previous_date,
            current_date,
        ),
        subtype="html",
    )

    smtp_send_message(
        message,
        settings,
    )

    mark_email_as_sent(
        signature,
        current_date,
    )

    return True


# -----------------------------------------------------------------------------
# Browser / execução
# -----------------------------------------------------------------------------


def create_browser_context(browser: Browser) -> BrowserContext:
    """
    Cria o contexto do navegador.

    O if separado evita um dict heterogêneo em **kwargs e mantém
    a análise do Pylance mais previsível.
    """
    if SESSION_FILE.exists():
        return browser.new_context(
            viewport={
                "width": 1600,
                "height": 1000,
            },
            locale="pt-BR",
            storage_state=str(SESSION_FILE),
        )

    return browser.new_context(
        viewport={
            "width": 1600,
            "height": 1000,
        },
        locale="pt-BR",
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=("Monitora a Carteira Finclass e mantém histórico diário.")
    )

    mode_group = parser.add_mutually_exclusive_group()

    mode_group.add_argument(
        "--test-email",
        action="store_true",
        help=(
            "Testa apenas o SMTP e encerra, "
            "sem acessar a Finclass ou alterar o histórico."
        ),
    )

    mode_group.add_argument(
        "--simulate-change",
        action="store_true",
        help=(
            "Envia uma notificação de alteração simulada usando "
            "o mesmo layout da produção, sem acessar a Finclass "
            "e sem alterar o histórico."
        ),
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()

    if args.test_email:
        send_test_email()
        return

    if args.simulate_change:
        send_simulated_change_email()
        return

    email, password = get_credentials()
    today = date.today()

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            headless=HEADLESS,
        )

        context = create_browser_context(browser)
        page = context.new_page()

        page.set_default_timeout(DEFAULT_TIMEOUT_MS)

        try:
            ensure_logged_in(
                page,
                context,
                email,
                password,
            )

            open_wallet(page)

            tables = scrape_all_recommendation_tables(page)

            total_rows = sum(len(table["rows"]) for table in tables)

            if total_rows == 0:
                raise RuntimeError(
                    "Nenhum item foi encontrado nas recommendation-table."
                )

            current_output = export_excel(tables)

            (
                assets_output,
                assets_total,
                unique_assets_total,
            ) = export_asset_catalog(tables)

            if assets_total != total_rows:
                raise RuntimeError(
                    "Falha de consistência na lista de ativos: foram coletados "
                    f"{total_rows} item(ns), mas o arquivo de ativos recebeu {assets_total}."
                )

            (
                history_output,
                previous_date,
                changes,
                snapshot_items,
            ) = save_daily_history(
                tables,
                today,
            )

            if snapshot_items != total_rows:
                raise RuntimeError(
                    f"Falha de consistência: foram coletados {total_rows} item(ns), "
                    f"mas o snapshot possui {snapshot_items}. O histórico não pode ser considerado completo."
                )

            print()
            print("Concluído.")
            print(f"Categorias coletadas: {len(tables)}")
            print(f"Itens coletados: {total_rows}")
            print(f"Snapshot histórico: {snapshot_items} itens")
            print(f"Carteira atual: {current_output}")
            print(f"Lista de ativos: {assets_output}")
            print(f"Itens processados na lista: {assets_total}")
            print(f"Nomes únicos: {unique_assets_total}")
            print(f"Histórico: {history_output}")

            if previous_date is None:
                print(
                    "Histórico: base inicial criada em {}.".format(
                        today.strftime("%d/%m/%Y")
                    )
                )
            else:
                print(
                    "Comparação: {} x {}".format(
                        previous_date.strftime("%d/%m/%Y"),
                        today.strftime("%d/%m/%Y"),
                    )
                )
                print(f"Alterações detectadas: {len(changes)}")

                for change in changes:
                    print(
                        "  - {} | {} | {} | {} -> {}".format(
                            change["change_type"],
                            change["ticker"] or change["company"],
                            change["company"],
                            percentage_text(change["previous_percentage"]),
                            percentage_text(change["current_percentage"]),
                        )
                    )

                if changes:
                    email_sent = send_changes_email(
                        changes,
                        previous_date,
                        today,
                    )

                    if email_sent:
                        print(
                            "E-mail: notificação enviada para {}.".format(
                                required_env("EMAIL_TO")
                            )
                        )
                else:
                    print("E-mail: nenhuma alteração; nenhum e-mail enviado.")

            print(
                "Executado em: {}".format(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
            )

        finally:
            context.close()
            browser.close()


if __name__ == "__main__":
    main()
